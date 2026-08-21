#!/usr/bin/env python3
"""Generate site content from the product spreadsheet.

`product-data.xlsx` (repo root) is the single source of truth. Its `Products`
and `Site Settings` sheets drive:
  * products/<slug>.html          — one detail page per model
  * index.html   #modelGrid       — the homepage model cards (CARDS markers)
  * app.js       COMPARE_MODELS    — data for the compare tray/modal/table

Run from the repo root:  python3 scripts/generate_products.py
Requires: openpyxl  (pip install openpyxl)
"""
import os
import re
from urllib.parse import quote

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_PATH = os.path.join(ROOT, "product-data.xlsx")

BADGES = {
    "in": ("badge-in", "In stock"),
    "limited": ("badge-limited", "Limited stock"),
    "pre": ("badge-pre", "Pre-order"),
}
STATUS_TO_BADGE = {
    "in stock": "in",
    "limited stock": "limited",
    "pre-order": "pre",
    "preorder": "pre",
}
# Card media background tint. Keeps the current homepage look; unknown slugs rotate.
MEDIA_BY_SLUG = {
    "glide-s1": "media-teal", "cruise-c3": "media-teal",
    "urban-u2": "media-indigo", "terra-x": "media-indigo",
    "compact-air": "media-amber", "recline-r5": "media-amber",
}
MEDIA_ROTATION = ["media-teal", "media-indigo", "media-amber"]

# These globals are filled from the Site Settings sheet in load_data().
WA = "923001234567"
PHONE = "+92 300 1234567"
HOURS = "Sat–Thu, 10am–7pm"
GENERAL_WA = ""


# --------------------------------------------------------------------------- #
# Load
# --------------------------------------------------------------------------- #
def _clean(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def load_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- Site settings ----
    settings = {}
    if "Site Settings" in wb.sheetnames:
        for row in wb["Site Settings"].iter_rows(values_only=True):
            if row and row[0] and len(row) > 1 and row[1] is not None:
                settings[_clean(row[0])] = _clean(row[1])

    # ---- Products ----
    ws = wb["Products"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 0 is the group banner, row 1 is the real header, data starts at row 2.
    header = [_clean(c) for c in rows[1]]
    idx = {name: i for i, name in enumerate(header)}

    def cell(row, name):
        i = idx.get(name)
        return _clean(row[i]) if i is not None and i < len(row) else ""

    products = []
    for row in rows[2:]:
        if not row or not _clean(row[idx["Product Name"]]):
            continue
        name = cell(row, "Product Name")
        slug = cell(row, "Slug (URL id)")
        badge = STATUS_TO_BADGE.get(cell(row, "Stock Status").lower(), "in")
        price_num = cell(row, "Price (PKR)")
        price = f"PKR {int(price_num):,}" if price_num.isdigit() else price_num

        rng = cell(row, "Range (km)")
        spd = cell(row, "Top Speed (km/h)")
        load = cell(row, "Max Load (kg)")
        kerb = cell(row, "Kerb Weight (kg)")

        # 4th card tile is author-controlled (label + value, value carries its own unit)
        quick = [
            ("Range", f"{rng} km"),
            ("Top speed", f"{spd} km/h"),
            ("Max load", f"{load} kg"),
            (cell(row, "Card Spec 4 — Label"), cell(row, "Card Spec 4 — Value")),
        ]

        # Full spec table: fixed order, blanks dropped.
        spec_src = [
            ("Range per charge", f"{rng} km" if rng else ""),
            ("Top speed", f"{spd} km/h" if spd else ""),
            ("Max load", f"{load} kg" if load else ""),
            ("Kerb weight", f"{kerb} kg" if kerb else ""),
            ("Battery", cell(row, "Battery")),
            ("Motor", cell(row, "Motor")),
            ("Charge time", cell(row, "Charge Time")),
            ("Foldable", cell(row, "Foldable")),
            ("Suspension", cell(row, "Suspension")),
            ("Seat", cell(row, "Seat")),
            ("Backrest", cell(row, "Backrest")),
            ("Legrests", cell(row, "Legrests")),
            ("Tyres", cell(row, "Tyres")),
            ("Controls", cell(row, "Controls")),
            ("Warranty", cell(row, "Warranty")),
        ]
        specs = [(k, v) for k, v in spec_src if v]

        highlights = [
            cell(row, f"Highlight {n}") for n in range(1, 6) if cell(row, f"Highlight {n}")
        ]

        # Compare-view specs (app.js). Missing → em dash.
        backrest, legrests = cell(row, "Backrest"), cell(row, "Legrests")
        if backrest and legrests:
            recline = f"{backrest.replace(' recline', '')} + {legrests.lower()} legrests"
        elif backrest:
            recline = backrest
        else:
            recline = "—"
        compare_specs = [
            ("Range", f"{rng} km"),
            ("Top speed", f"{spd} km/h"),
            ("Max load", f"{load} kg"),
            ("Kerb weight", f"{kerb} kg"),
            ("Battery", cell(row, "Battery") or "—"),
            ("Motor", cell(row, "Motor") or "—"),
            ("Charge time", cell(row, "Charge Time") or "—"),
            ("Foldable", cell(row, "Foldable") or "—"),
            ("Suspension", cell(row, "Suspension") or "—"),
            ("Tyres", cell(row, "Tyres") or "—"),
            ("Controls", cell(row, "Controls") or "—"),
            ("Recline", recline),
            ("Warranty", cell(row, "Warranty") or "—"),
        ]

        lead = cell(row, "Lead Paragraph")
        products.append({
            "slug": slug,
            "name": name,
            "category": cell(row, "Category"),
            "tag": cell(row, "Tag / Label"),
            "badge": badge,
            "price": price,
            "image": cell(row, "Image File") or "q5.png",
            "media": MEDIA_BY_SLUG.get(slug, MEDIA_ROTATION[len(products) % 3]),
            "card_desc": cell(row, "Card Description"),
            "lead": lead,
            "meta": cell(row, "SEO Meta Description") or lead[:150],
            "quick": quick,
            "specs": specs,
            "compare_specs": compare_specs,
            "highlights": highlights,
            "bestfor": cell(row, "Best For"),
        })
    return products, settings


# --------------------------------------------------------------------------- #
# WhatsApp links
# --------------------------------------------------------------------------- #
def wa_link(p):
    if p["badge"] == "pre":
        msg = f"Hi Stride, I'd like to pre-order the {p['name']}. Please share details and timeline."
    else:
        msg = (f"Hi Stride, I'm interested in the {p['name']} electric wheelchair. "
               "Please share availability and final price.")
    return f"https://wa.me/{WA}?text={quote(msg)}"


def wa_link_short(p):
    """Homepage-card variant (no 'electric wheelchair'), matching the existing copy."""
    if p["badge"] == "pre":
        msg = f"Hi Stride, I'd like to pre-order the {p['name']}. Please share details and timeline."
    else:
        msg = f"Hi Stride, I'm interested in the {p['name']}. Please share availability and final price."
    return f"https://wa.me/{WA}?text={quote(msg)}"


def related_products(products, p, limit=3):
    """Same-category models first, then others; excludes the current product."""
    rest = [x for x in products if x["slug"] != p["slug"]]
    same = [x for x in rest if x["category"] == p["category"]]
    others = [x for x in rest if x["category"] != p["category"]]
    return (same + others)[:limit]


# --------------------------------------------------------------------------- #
# Homepage cards + compare data
# --------------------------------------------------------------------------- #
def home_card_html(p):
    badge_class, badge_text = BADGES[p["badge"]]
    specs = "\n".join(
        f"                <li><span>{k}</span><strong>{v}</strong></li>" for k, v in p["quick"]
    )
    return f"""          <article class="card" data-category="{p['category']}">
            <div class="card-media {p['media']}">
              <span class="badge {badge_class}">{badge_text}</span>
              <button class="compare-toggle" type="button" data-compare="{p['slug']}" aria-pressed="false">
                <svg class="icon" aria-hidden="true"><use href="#ic-check"/></svg><span class="compare-toggle-label">Compare</span>
              </button>
              <img class="card-photo" src="images/{p['image']}" alt="{p['name']} carbon-fibre electric wheelchair" loading="lazy">
            </div>
            <div class="card-body">
              <div class="card-top">
                <h3><a class="card-title-link" href="products/{p['slug']}.html">{p['name']}</a></h3>
                <span class="tag">{p['tag']}</span>
              </div>
              <p class="card-desc">{p['card_desc']}</p>
              <ul class="specs">
{specs}
              </ul>
              <div class="card-foot">
                <div class="price">{p['price']}</div>
                <a class="btn btn-primary btn-block" href="tel:+{WA}">
                  <svg class="icon" aria-hidden="true"><use href="#ic-phone"/></svg><span>Call now</span>
                </a>
                <a class="btn btn-wa btn-block" href="{wa_link_short(p)}" target="_blank" rel="noopener">
                  <svg class="icon" aria-hidden="true"><use href="#ic-wa"/></svg><span>Chat on WhatsApp</span>
                </a>
              </div>
            </div>
          </article>"""


def _js_str(s):
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'") + "'"


def compare_models_block(products):
    lines = ["const COMPARE_MODELS = {"]
    for p in products:
        specs = ", ".join(f"{_js_str(k)}: {_js_str(v)}" for k, v in p["compare_specs"])
        lines.append(f"  {_js_str(p['slug'])}: {{")
        lines.append(
            f"    name: {_js_str(p['name'])}, tag: {_js_str(p['tag'])}, "
            f"price: {_js_str(p['price'])}, img: {_js_str('images/' + p['image'])},"
        )
        lines.append(f"    wa: {_js_str(wa_link_short(p))},")
        lines.append(f"    specs: {{ {specs} }},")
        lines.append("  },")
    lines.append("};")
    return "\n".join(lines)


def inject(path, start_marker, end_marker, content):
    """Replace whatever sits between two markers (markers kept) with `content`."""
    with open(path) as f:
        text = f.read()
    pattern = re.compile(
        re.escape(start_marker) + r".*?" + re.escape(end_marker), re.DOTALL
    )
    if not pattern.search(text):
        raise SystemExit(f"markers not found in {os.path.relpath(path, ROOT)}")
    replacement = f"{start_marker}\n{content}\n{end_marker}"
    text = pattern.sub(lambda _m: replacement, text, count=1)
    with open(path, "w") as f:
        f.write(text)


SYMBOLS = """  <svg width="0" height="0" style="position:absolute" aria-hidden="true">
    <symbol id="ic-wheelchair" viewBox="0 0 24 24">
      <circle cx="8.5" cy="18" r="3.2" fill="none" stroke="currentColor" stroke-width="1.4"/>
      <circle cx="8.5" cy="18" r="0.6" fill="currentColor"/>
      <circle cx="6.5" cy="6" r="1.8" fill="currentColor"/>
      <path d="M6.5 8.4v4.2h4.4l2.6 4.6" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"/>
      <path d="M6.9 10.4h4" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round"/>
      <path d="M13.5 17.2h3.2l1.8-4" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round"/>
    </symbol>
    <symbol id="ic-wa" viewBox="0 0 24 24">
      <path d="M12 2a10 10 0 0 0-8.6 15.1L2 22l5.1-1.3A10 10 0 1 0 12 2Zm5.4 14.3c-.2.6-1.3 1.2-1.8 1.2-.5.1-1 .2-3.3-.7-2.8-1.1-4.5-3.9-4.7-4.1-.1-.2-1.1-1.4-1.1-2.7 0-1.3.7-1.9.9-2.2.2-.2.5-.3.7-.3h.5c.2 0 .4 0 .6.5l.8 1.9c.1.2.1.4 0 .5l-.4.6c-.2.2-.3.4-.1.7.2.3.9 1.4 1.9 2.3 1.3 1.1 2.3 1.5 2.6 1.6.3.1.5.1.7-.1l.9-1c.2-.3.4-.2.7-.1l1.9.9c.3.1.5.2.6.3.1.3.1.8-.1 1.4Z" fill="currentColor"/>
    </symbol>
    <symbol id="ic-check" viewBox="0 0 24 24">
      <path d="M20 6 9 17l-5-5" fill="none" stroke="currentColor" stroke-width="2.4" stroke-linecap="round" stroke-linejoin="round"/>
    </symbol>
    <symbol id="ic-truck" viewBox="0 0 24 24">
      <path d="M2.5 6h11v9h-11z" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linejoin="round"/>
      <path d="M13.5 9h4l3.5 3.2V15h-7.5z" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linejoin="round"/>
      <circle cx="7" cy="17.5" r="1.9" fill="none" stroke="currentColor" stroke-width="1.6"/>
      <circle cx="17.5" cy="17.5" r="1.9" fill="none" stroke="currentColor" stroke-width="1.6"/>
    </symbol>
    <symbol id="ic-phone" viewBox="0 0 24 24">
      <path d="M6.6 10.8c1.4 2.8 3.8 5.1 6.6 6.6l2.2-2.2c.3-.3.7-.4 1-.2 1.1.4 2.3.6 3.6.6.6 0 1 .4 1 1V20c0 .6-.4 1-1 1-9.4 0-17-7.6-17-17 0-.6.4-1 1-1h3.5c.6 0 1 .4 1 1 0 1.3.2 2.5.6 3.6.1.4 0 .7-.2 1z" fill="currentColor"/>
    </symbol>
  </svg>"""


# --------------------------------------------------------------------------- #
# Detail-page card + page
# --------------------------------------------------------------------------- #
def card_html(rp):
    """A compact model card (matches the homepage grid) for the related section."""
    badge_class, badge_text = BADGES[rp["badge"]]
    short = (rp["lead"].split(". ")[0].rstrip(".") + ".") if rp["lead"] else rp["card_desc"]
    specs = "\n".join(
        f'                <li><span>{k}</span><strong>{v}</strong></li>' for k, v in rp["quick"]
    )
    return f"""            <article class="card">
              <div class="card-media {rp['media']}">
                <span class="badge {badge_class}">{badge_text}</span>
                <svg class="card-chair"><use href="#ic-wheelchair"/></svg>
              </div>
              <div class="card-body">
                <div class="card-top">
                  <h3><a class="card-title-link" href="{rp['slug']}.html">{rp['name']}</a></h3>
                  <span class="tag">{rp['tag']}</span>
                </div>
                <p class="card-desc">{short}</p>
                <ul class="specs">
{specs}
                </ul>
                <div class="card-foot">
                  <div class="price">{rp['price']}</div>
                  <a class="btn btn-primary btn-block" href="tel:+{WA}">
                    <svg class="icon" aria-hidden="true"><use href="#ic-phone"/></svg><span>Call now</span>
                  </a>
                  <a class="btn btn-wa btn-block" href="{wa_link(rp)}" target="_blank" rel="noopener">
                    <svg class="icon" aria-hidden="true"><use href="#ic-wa"/></svg><span>Chat on WhatsApp</span>
                  </a>
                </div>
              </div>
            </article>"""


def render(p, products):
    badge_class, badge_text = BADGES[p["badge"]]
    quick = "\n".join(
        f'              <li><span>{k}</span><strong>{v}</strong></li>' for k, v in p["quick"]
    )
    specs = "\n".join(
        f'              <tr><th>{k}</th><td>{v}</td></tr>' for k, v in p["specs"]
    )
    highlights = "\n".join(
        f'              <li><svg class="icon" aria-hidden="true"><use href="#ic-check"/></svg><span>{h}</span></li>'
        for h in p["highlights"]
    )
    related_html = "\n".join(card_html(rp) for rp in related_products(products, p))
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{p['name']} Stride Electric Wheelchairs</title>
  <meta name="description" content="{p['meta']}">
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Google+Sans+Flex:wght@400..800&display=swap">
  <link rel="stylesheet" href="../style.css">
</head>
<body>

{SYMBOLS}

  <div class="topbar" id="top">
    <div class="container topbar-inner">
      <span class="topbar-msg"><svg class="icon" aria-hidden="true"><use href="#ic-truck"/></svg> Free delivery across Pakistan</span>
      <span class="topbar-contact">
        <a href="{GENERAL_WA}" target="_blank" rel="noopener"><svg class="icon" aria-hidden="true"><use href="#ic-wa"/></svg> {PHONE}</a>
        <span class="topbar-sep" aria-hidden="true">·</span>
        <span class="topbar-hours">{HOURS}</span>
      </span>
    </div>
  </div>
  <header class="site-header" id="headerMain">
    <div class="container header-inner">
      <a href="../index.html" class="brand" aria-label="Stride home">
        <img src="../images/logo.svg" alt="Stride" class="brand-logo">
      </a>
      <nav class="nav" aria-label="Primary">
        <a href="../index.html#models">Models</a>
        <a href="../index.html#why">Why Stride</a>
        <a href="../index.html#compare">Compare</a>
        <a href="../index.html#contact">Contact</a>
      </nav>
      <a class="btn btn-primary header-cta" href="tel:+{WA}">
        <svg class="icon" aria-hidden="true"><use href="#ic-phone"/></svg>
        <span>Call now</span>
      </a>
      <button class="nav-toggle" id="navToggle" aria-label="Toggle menu" aria-expanded="false">
        <span></span><span></span><span></span>
      </button>
    </div>
    <nav class="mobile-nav" id="mobileNav" aria-label="Mobile">
      <a href="../index.html#models">Models</a>
      <a href="../index.html#why">Why Stride</a>
      <a href="../index.html#compare">Compare</a>
      <a href="../index.html#contact">Contact</a>
      <a class="btn btn-primary" href="tel:+{WA}">Call now for free consultation</a>
      <a class="btn btn-wa" href="{wa_link(p)}" target="_blank" rel="noopener">Chat on WhatsApp</a>
    </nav>
  </header>

  <main>
    <section class="section pdp">
      <div class="container">
        <a class="back-link" href="../index.html#models">← All models</a>

        <div class="pdp-hero">
          <div class="pdp-media {p['media']}">
            <span class="badge {badge_class}">{badge_text}</span>
            <svg class="pdp-chair"><use href="#ic-wheelchair"/></svg>
          </div>
          <div class="pdp-info">
            <span class="tag">{p['tag']}</span>
            <h1>{p['name']}</h1>
            <p class="pdp-lead">{p['lead']}</p>
            <div class="pdp-price">{p['price']}</div>
            <a class="btn btn-primary btn-lg" href="tel:+{WA}">
              <svg class="icon" aria-hidden="true"><use href="#ic-phone"/></svg><span>Call now for free consultation</span>
            </a>
            <a class="btn btn-wa btn-lg" href="{wa_link(p)}" target="_blank" rel="noopener">
              <svg class="icon" aria-hidden="true"><use href="#ic-wa"/></svg><span>Chat on WhatsApp</span>
            </a>
            <ul class="pdp-quick">
{quick}
            </ul>
          </div>
        </div>

        <div class="pdp-grid">
          <section class="pdp-block">
            <h2>Full specifications</h2>
            <table class="spec-table">
{specs}
            </table>
          </section>
          <section class="pdp-block">
            <h2>Highlights</h2>
            <ul class="highlight-list">
{highlights}
            </ul>
            <div class="bestfor"><strong>Best for:</strong> {p['bestfor']}</div>
          </section>
        </div>
      </div>
    </section>

    <section class="section section-alt">
      <div class="container">
        <div class="section-head">
          <h2>Related models</h2>
          <p>Other Stride chairs worth comparing before you decide.</p>
        </div>
        <div class="grid">
{related_html}
        </div>
      </div>
    </section>

    <section class="section contact" id="contact">
      <div class="container">
        <div class="contact-card">
          <div class="contact-copy">
            <h2>Questions about the {p['name']}?</h2>
            <p>Call us for a free consultation on live stock, delivery time to your city and the final quote. Prefer to type? Message us on WhatsApp and we'll send real photos and answer anything.</p>
            <div class="contact-meta">
              <p><strong>WhatsApp / Phone:</strong> {PHONE}</p>
              <p><strong>Hours:</strong> {HOURS} · Serving all of Pakistan</p>
            </div>
          </div>
          <div class="contact-cta">
            <a class="btn btn-primary btn-lg" href="tel:+{WA}">
              <svg class="icon" aria-hidden="true"><use href="#ic-phone"/></svg>
              <span>Call now for free consultation</span>
            </a>
            <a class="btn btn-wa btn-lg" href="{wa_link(p)}" target="_blank" rel="noopener">
              <svg class="icon" aria-hidden="true"><use href="#ic-wa"/></svg>
              <span>Chat on WhatsApp</span>
            </a>
          </div>
        </div>
      </div>
    </section>
  </main>

  <footer class="site-footer">
    <div class="container footer-inner">
      <div class="footer-brand">
        <img src="../images/logo-light.svg" alt="Stride" class="brand-logo footer-logo">
      </div>
      <p class="footer-tag">Electric wheelchairs for independent living, delivered across Pakistan.</p>
      <nav class="footer-nav" aria-label="Footer">
        <a href="../index.html#models">Models</a>
        <a href="../index.html#why">Why Stride</a>
        <a href="../index.html#contact">Contact</a>
      </nav>
    </div>
    <div class="container footer-bottom">
      <span>© <span id="year">2026</span> Stride. All rights reserved.</span>
      <span>Prices in PKR and subject to change. Contact us for current pricing.</span>
    </div>
  </footer>

  <a class="call-float" href="tel:+{WA}" aria-label="Call now for a free consultation">
    <svg class="icon" aria-hidden="true"><use href="#ic-phone"/></svg>
  </a>
  <a class="wa-float" href="{wa_link(p)}" target="_blank" rel="noopener" aria-label="Chat on WhatsApp">
    <svg class="icon" aria-hidden="true"><use href="#ic-wa"/></svg>
  </a>

  <script src="../app.js"></script>
</body>
</html>
"""


def main():
    global WA, PHONE, HOURS, GENERAL_WA
    products, settings = load_data(DATA_PATH)
    WA = settings.get("WhatsApp number (digits, for links)", WA)
    PHONE = settings.get("Business phone (display)", PHONE)
    HOURS = settings.get("Business hours", HOURS)
    GENERAL_WA = f"https://wa.me/{WA}?text=" + quote(
        "Hi Stride, I'd like to know more about your electric wheelchairs."
    )

    # 1) Detail pages
    out_dir = os.path.join(ROOT, "products")
    os.makedirs(out_dir, exist_ok=True)
    for p in products:
        path = os.path.join(out_dir, f"{p['slug']}.html")
        with open(path, "w") as f:
            f.write(render(p, products))
        print(f"wrote products/{p['slug']}.html")

    # 2) Homepage cards
    cards = "\n".join(home_card_html(p) for p in products)
    inject(
        os.path.join(ROOT, "index.html"),
        "<!-- CARDS:START — generated from product-data.xlsx by scripts/generate_products.py -->",
        "<!-- CARDS:END -->",
        "\n" + cards + "\n",
    )
    print("updated index.html #modelGrid")

    # 3) Compare data in app.js
    inject(
        os.path.join(ROOT, "app.js"),
        "/* COMPARE_MODELS:START — generated from product-data.xlsx by scripts/generate_products.py */",
        "/* COMPARE_MODELS:END */",
        compare_models_block(products),
    )
    print("updated app.js COMPARE_MODELS")


if __name__ == "__main__":
    main()
